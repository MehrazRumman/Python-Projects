#! python3
# update Produce- corrects costs in produce sales spreadsheet
import openpyxl
wb=openpyxl.load_workbook("d:\\produceSales.xlsx")
sheet=wb["Sheet"]
# the produce type and their update prices
update={"Garlic":3.07,"Celery":1.19,"Lemon":1.27}

# TODO: Loop through the rows and update the prices
for row in range(2,sheet.max_row):
    pn=sheet.cell(row=row,column=1).value
    if pn in update:
        sheet.cell(row=row,column=2).value=update[pn]
wb.save("d:\\updatedproducesales.xlsx")